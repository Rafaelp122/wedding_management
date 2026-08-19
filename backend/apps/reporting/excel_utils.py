"""
Utilitários, estilos e renderizador openpyxl para planilhas Excel (DESIGN.md).
"""

import io
from datetime import UTC, datetime
from decimal import Decimal
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from apps.finances.models import BudgetCategory, Installment
from apps.logistics.models import Contract
from apps.scheduler.models import Task


def _build_excel_summary_sheet(
    wb: Workbook,
    wedding: Any,
    overview: dict[str, Any],
    installments: list[Installment],
    tasks: list[Task],
) -> Worksheet:
    """Constrói a aba de Resumo Executivo da planilha alinhada ao DESIGN.md."""
    ws_summary = cast(Worksheet, wb.active)
    ws_summary.title = "Resumo Executivo"

    title_font = Font(name="Calibri", size=14, bold=True, color="7C3AED")
    regular_font = Font(name="Calibri", size=11, color="1A1C1E")

    now_label = datetime.now(UTC).strftime("%d/%m/%Y às %H:%M UTC")
    ws_summary.cell(
        row=1,
        column=1,
        value=f"Relatório: {wedding.groom_name} & {wedding.bride_name}",
    ).font = title_font
    ws_summary.cell(
        row=2,
        column=1,
        value=f"Emitido em: {now_label} (Sim, Aceito! Prestige)",
    ).font = regular_font

    ws_summary.append([])
    ws_summary.append(["Informações Gerais", "Valor"])
    ws_summary.append(["Noivo", wedding.groom_name])
    ws_summary.append(["Noiva", wedding.bride_name])
    wedding_date_str = wedding.date.strftime("%d/%m/%Y") if wedding.date else "—"
    ws_summary.append(["Data do Casamento", wedding_date_str])
    ws_summary.append(["Local", wedding.location or "—"])
    ws_summary.append(["Convidados Estimados", wedding.expected_guests or "—"])
    ws_summary.append(["Status do Casamento", wedding.get_status_display()])

    ws_summary.append([])
    ws_summary.append(["Métrica Financeira", "Valor"])
    budget_obj = getattr(wedding, "budget", None)
    total_budget_val = float(budget_obj.total_estimated) if budget_obj else 0.0
    ws_summary.append(["Orçamento Total", total_budget_val])
    paid_sum = sum(
        (i.amount for i in installments if i.status == Installment.StatusChoices.PAID),
        Decimal("0.00"),
    )
    ws_summary.append(["Total Pago", float(paid_sum)])

    pending_sum = sum(
        (
            i.amount
            for i in installments
            if i.status
            in (
                Installment.StatusChoices.PENDING,
                Installment.StatusChoices.OVERDUE,
            )
        ),
        Decimal("0.00"),
    )
    ws_summary.append(["Total Pendente / Atrasado", float(pending_sum)])

    budget_pct_used = overview.get("budget_percentage_used", 0)
    ws_summary.append(["Saúde Financeira Utilizada (%)", f"{budget_pct_used}%"])
    ws_summary.append(
        [
            "Tarefas Concluídas",
            f"{sum(1 for t in tasks if t.is_completed)} de {len(tasks)}",
        ]
    )
    return ws_summary


def _build_excel_categories_sheet(
    wb: Workbook, categories: list[BudgetCategory]
) -> None:
    """Constrói a aba de Categorias Orçamentárias."""
    ws_cat = wb.create_sheet(title="Categorias Orçamentárias")
    ws_cat.append(
        [
            "Categoria",
            "Verba Alocada (R$)",
            "Total Gasto (R$)",
            "Saldo (R$)",
            "Uso (%)",
        ]
    )
    for cat in categories:
        spent = float(cat.total_spent)
        allocated = float(cat.allocated_budget)
        remaining = allocated - spent
        pct = round((spent / allocated) * 100, 1) if allocated > 0 else 0
        ws_cat.append([cat.name, allocated, spent, remaining, f"{pct}%"])


def _build_excel_installments_sheet(
    wb: Workbook, installments: list[Installment]
) -> None:
    """Constrói a aba de Cronograma de Parcelas."""
    ws_inst = wb.create_sheet(title="Cronograma de Parcelas")
    ws_inst.append(
        [
            "Despesa / Item",
            "Parcela Nº",
            "Vencimento",
            "Valor (R$)",
            "Status",
            "Data de Pagamento",
        ]
    )
    for inst in installments:
        desc = inst.expense.description if inst.expense else "Parcela"
        due_str = inst.due_date.strftime("%d/%m/%Y")
        paid_str = inst.paid_date.strftime("%d/%m/%Y") if inst.paid_date else "—"
        ws_inst.append(
            [
                desc,
                inst.installment_number,
                due_str,
                float(inst.amount),
                inst.get_status_display(),
                paid_str,
            ]
        )


def _build_excel_contracts_sheet(wb: Workbook, contracts: list[Contract]) -> None:
    """Constrói a aba de Contratos & Fornecedores."""
    ws_contr = wb.create_sheet(title="Contratos & Fornecedores")
    ws_contr.append(
        [
            "Fornecedor",
            "Nome / Descrição",
            "Valor Total (R$)",
            "Status",
            "Data de Expiração",
        ]
    )
    for c in contracts:
        sup_name = c.supplier.name if c.supplier else "Fornecedor Direto"
        exp_str = c.expiration_date.strftime("%d/%m/%Y") if c.expiration_date else "—"
        ws_contr.append(
            [
                sup_name,
                c.name,
                float(c.total_amount),
                c.get_status_display(),
                exp_str,
            ]
        )


def _build_excel_tasks_sheet(wb: Workbook, tasks: list[Task]) -> None:
    """Constrói a aba de Checklist de Tarefas."""
    ws_tasks = wb.create_sheet(title="Checklist de Tarefas")
    ws_tasks.append(["Status", "Título da Tarefa", "Prazo", "Descrição"])
    for t in tasks:
        status_sym = "Concluída" if t.is_completed else "Pendente"
        due_str = t.due_date.strftime("%d/%m/%Y") if t.due_date else "—"
        ws_tasks.append([status_sym, t.title, due_str, t.description])


def _format_columns_and_borders(sheet: Worksheet) -> None:
    """Aplica bordas finas, formatação monetária e auto-ajuste de colunas."""
    thin_border = Border(
        left=Side(style="thin", color="E4E4E7"),
        right=Side(style="thin", color="E4E4E7"),
        top=Side(style="thin", color="E4E4E7"),
        bottom=Side(style="thin", color="E4E4E7"),
    )
    currency_format = '"R$" #,##0.00'
    monetary_keys = ["(r$)", "valor", "orçamento", "pago", "gasto", "saldo"]

    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(cast(int, col[0].column))
        for cell in col:
            cell.border = thin_border
            col_idx = cast(int, cell.column)
            header_val = str(sheet.cell(row=1, column=col_idx).value).lower()
            if isinstance(cell.value, (int, float)) and any(
                k in header_val for k in monetary_keys
            ):
                cell.number_format = currency_format
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)


def _style_excel_workbook(wb: Workbook, ws_summary: Worksheet) -> None:
    """Aplica cores do DESIGN.md (Prestige Purple #7C3AED) e bordas #E4E4E7."""
    header_fill = PatternFill(
        start_color="7C3AED", end_color="7C3AED", fill_type="solid"
    )
    section_fill = PatternFill(
        start_color="F5F3FF", end_color="F5F3FF", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=11, bold=True, color="7C3AED")

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            if sheet != ws_summary:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        if sheet == ws_summary:
            for row_num in (4, 12):
                for cell in sheet[row_num]:
                    cell.fill = section_fill
                    cell.font = section_font

        _format_columns_and_borders(sheet)


def render_wedding_excel(
    wedding: Any,
    overview: dict[str, Any],
    categories: list[BudgetCategory],
    installments: list[Installment],
    contracts: list[Contract],
    tasks: list[Task],
) -> bytes:
    """
    Renderiza a planilha Excel (.xlsx) completa com 5 abas estilizadas (DESIGN.md).
    """
    wb = Workbook()
    ws_summary = _build_excel_summary_sheet(wb, wedding, overview, installments, tasks)
    _build_excel_categories_sheet(wb, categories)
    _build_excel_installments_sheet(wb, installments)
    _build_excel_contracts_sheet(wb, contracts)
    _build_excel_tasks_sheet(wb, tasks)
    _style_excel_workbook(wb, ws_summary)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
