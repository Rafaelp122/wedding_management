"""
Camada de serviços para o módulo de reporting (relatórios, exportações e métricas).
"""

import io
from datetime import UTC, datetime
from decimal import Decimal
from typing import Any, Literal, cast
from uuid import UUID

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from apps.core.services import get_storage_service
from apps.core.services.storage import StorageService
from apps.finances.models import BudgetCategory, Installment
from apps.logistics.models import Contract
from apps.reporting.pdf_utils import NumberedCanvas
from apps.reporting.selectors import wedding_overview_selector
from apps.scheduler.models import Task
from apps.tenants.models import Company
from apps.weddings.selectors import wedding_get_selector


class ReportGenerationService:
    """
    Serviço de geração e exportação de relatórios consolidados do casamento.

    Consolida informações multi-tenant dos domínios de casamentos, finanças,
    logística e cronograma em arquivos binários diagramados (PDF e Excel).
    """

    _storage_service: StorageService | None = None

    @classmethod
    def _get_storage_service(cls) -> StorageService:
        """
        Retorna a instância do serviço de storage injetado ou inicializado.

        Returns:
            Instância ativa de StorageService.
        """
        if cls._storage_service is None:
            cls._storage_service = get_storage_service()
        return cls._storage_service

    @classmethod
    def _set_storage_service(cls, storage_service: StorageService | None) -> None:
        """
        Injeta uma instância customizada de StorageService para testes.

        Args:
            storage_service: Instância mock ou customizada do protocolo.
        """
        cls._storage_service = storage_service

    @classmethod
    def _format_currency(cls, val: Decimal | float | int | None) -> str:
        """Formata valor decimal em moeda brasileira (R$ 1.234,56)."""
        if val is None:
            return "R$ 0,00"
        dec = Decimal(str(val))
        return f"R$ {dec:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    @classmethod
    def generate_wedding_pdf(
        cls,
        company: Company,
        wedding_uuid: UUID | str,
    ) -> bytes:
        """
        Gera um relatório diagramado completo do casamento em formato PDF.

        Args:
            company: Empresa tenant autenticada.
            wedding_uuid: Identificador único do casamento.

        Returns:
            Bytes do arquivo PDF gerado.
        """
        wedding = wedding_get_selector(
            company=company,
            uuid=UUID(str(wedding_uuid)),
        )
        overview = wedding_overview_selector(
            company=company,
            wedding_uuid=UUID(str(wedding_uuid)),
        )

        categories = list(
            BudgetCategory.objects.for_tenant(company)
            .filter(budget__wedding=wedding)
            .order_by("name")
        )
        installments = list(
            Installment.objects.for_tenant(company)
            .filter(wedding=wedding)
            .select_related("expense")
            .order_by("due_date")
        )
        contracts = list(
            Contract.objects.for_tenant(company)
            .filter(wedding=wedding)
            .select_related("supplier")
            .order_by("created_at")
        )
        tasks = list(
            Task.objects.for_tenant(company)
            .filter(wedding=wedding)
            .order_by("is_completed", "due_date")
        )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=2 * cm,
            rightMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2.2 * cm,
        )

        styles = getSampleStyleSheet()
        primary_color = colors.HexColor("#0F766E")
        dark_text = colors.HexColor("#18181B")
        muted_text = colors.HexColor("#71717A")
        border_color = colors.HexColor("#E4E4E7")
        bg_card = colors.HexColor("#F4F4F5")
        bg_alt = colors.HexColor("#FAFAFA")

        title_style = ParagraphStyle(
            "DocTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=dark_text,
            spaceAfter=4,
        )
        subtitle_style = ParagraphStyle(
            "DocSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=muted_text,
            spaceAfter=12,
        )
        section_style = ParagraphStyle(
            "SectionHeading",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=17,
            textColor=primary_color,
            spaceBefore=14,
            spaceAfter=6,
        )
        cell_style = ParagraphStyle(
            "TableCell",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            textColor=dark_text,
        )
        cell_header_style = ParagraphStyle(
            "TableHeader",
            parent=cell_style,
            fontName="Helvetica-Bold",
            textColor=colors.white,
        )

        story: list[Any] = []

        # 1. Cabeçalho Principal
        couple_names = f"{wedding.groom_name} & {wedding.bride_name}"
        story.append(Paragraph(couple_names, title_style))

        date_str = (
            wedding.date.strftime("%d/%m/%Y") if wedding.date else "Data não definida"
        )
        guests_str = (
            f"{wedding.expected_guests} convidados"
            if wedding.expected_guests
            else "Convidados não informados"
        )
        location_str = wedding.location or "Local não informado"
        status_label = wedding.get_status_display()

        header_info = (
            f"<b>Data:</b> {date_str} &nbsp;|&nbsp; "
            f"<b>Local:</b> {location_str} &nbsp;|&nbsp; "
            f"<b>Esperado:</b> {guests_str} &nbsp;|&nbsp; "
            f"<b>Status:</b> {status_label}"
        )
        story.append(Paragraph(header_info, subtitle_style))
        story.append(
            HRFlowable(
                width="100%",
                thickness=1,
                color=border_color,
                spaceBefore=0,
                spaceAfter=12,
            )
        )

        # 2. Cartões de Resumo Executivo / KPIs
        budget_obj = getattr(wedding, "budget", None)
        total_budget = budget_obj.total_estimated if budget_obj else Decimal("0.00")
        total_budget_str = cls._format_currency(total_budget)
        paid_installments_sum = sum(
            (
                i.amount
                for i in installments
                if i.status == Installment.StatusChoices.PAID
            ),
            Decimal("0.00"),
        )
        pending_installments_sum = sum(
            (
                i.amount
                for i in installments
                if i.status
                in (
                    Installment.StatusChoices.PENDING,
                    Installment.StatusChoices.OVERDUE,
                )
            ),
            Decimal("0.00"),
        )

        kpi_data = [
            [
                Paragraph("<b>ORÇAMENTO TOTAL</b>", subtitle_style),
                Paragraph("<b>TOTAL PAGO</b>", subtitle_style),
                Paragraph("<b>A PAGAR / PENDENTE</b>", subtitle_style),
                Paragraph("<b>SAÚDE FINANCEIRA</b>", subtitle_style),
            ],
            [
                Paragraph(f"<b>{total_budget_str}</b>", title_style),
                Paragraph(
                    f"<b>{cls._format_currency(paid_installments_sum)}</b>",
                    title_style,
                ),
                Paragraph(
                    f"<b>{cls._format_currency(pending_installments_sum)}</b>",
                    title_style,
                ),
                Paragraph(
                    f"<b>{overview.get('budget_percentage_used', 0)}%</b>",
                    title_style,
                ),
            ],
        ]
        kpi_table = Table(kpi_data, colWidths=[4.2 * cm, 4.2 * cm, 4.4 * cm, 4.2 * cm])
        kpi_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), bg_card),
                    ("BOX", (0, 0), (-1, -1), 0.5, border_color),
                    ("INNERGRID", (0, 0), (-1, -1), 0.5, border_color),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(kpi_table)
        story.append(Spacer(1, 10))

        # 3. Categorias Orçamentárias
        story.append(
            Paragraph("Distribuição por Categoria Orçamentária", section_style)
        )
        cat_data = [
            [
                Paragraph("Categoria", cell_header_style),
                Paragraph("Verba Alocada", cell_header_style),
                Paragraph("Total Gasto", cell_header_style),
                Paragraph("Saldo Restante", cell_header_style),
                Paragraph("Uso (%)", cell_header_style),
            ]
        ]
        for cat in categories:
            spent = cat.total_spent
            remaining = cat.allocated_budget - spent
            pct = (
                round((spent / cat.allocated_budget) * 100, 1)
                if cat.allocated_budget > 0
                else 0
            )
            cat_data.append(
                [
                    Paragraph(cat.name, cell_style),
                    Paragraph(cls._format_currency(cat.allocated_budget), cell_style),
                    Paragraph(cls._format_currency(spent), cell_style),
                    Paragraph(cls._format_currency(remaining), cell_style),
                    Paragraph(f"{pct}%", cell_style),
                ]
            )
        if len(cat_data) == 1:
            cat_data.append(
                [
                    Paragraph("Nenhuma categoria orçamentária cadastrada.", cell_style),
                    Paragraph("—", cell_style),
                    Paragraph("—", cell_style),
                    Paragraph("—", cell_style),
                    Paragraph("—", cell_style),
                ]
            )

        cat_table = Table(
            cat_data, colWidths=[5.5 * cm, 3.0 * cm, 3.0 * cm, 3.2 * cm, 2.3 * cm]
        )
        cat_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), primary_color),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
                    ("TOPPADDING", (0, 0), (-1, 0), 5),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_alt]),
                    ("GRID", (0, 0), (-1, -1), 0.5, border_color),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        story.append(cat_table)
        story.append(Spacer(1, 10))

        # 4. Cronograma de Parcelas & Vencimentos
        story.append(Paragraph("Cronograma de Parcelas & Pagamentos", section_style))
        inst_data = [
            [
                Paragraph("Despesa / Descrição", cell_header_style),
                Paragraph("Parcela", cell_header_style),
                Paragraph("Vencimento", cell_header_style),
                Paragraph("Valor", cell_header_style),
                Paragraph("Status", cell_header_style),
                Paragraph("Data Pagamento", cell_header_style),
            ]
        ]
        for inst in installments:
            due_str = inst.due_date.strftime("%d/%m/%Y")
            paid_str = inst.paid_date.strftime("%d/%m/%Y") if inst.paid_date else "—"
            status_desc = inst.get_status_display()
            desc = inst.expense.description if inst.expense else "Parcela"
            inst_data.append(
                [
                    Paragraph(desc, cell_style),
                    Paragraph(f"Nº {inst.installment_number}", cell_style),
                    Paragraph(due_str, cell_style),
                    Paragraph(cls._format_currency(inst.amount), cell_style),
                    Paragraph(status_desc, cell_style),
                    Paragraph(paid_str, cell_style),
                ]
            )
        if len(inst_data) == 1:
            inst_data.append(
                [
                    Paragraph("Nenhuma parcela cadastrada.", cell_style),
                    Paragraph("—", cell_style),
                    Paragraph("—", cell_style),
                    Paragraph("—", cell_style),
                    Paragraph("—", cell_style),
                    Paragraph("—", cell_style),
                ]
            )

        inst_table = Table(
            inst_data,
            colWidths=[
                5.0 * cm,
                2.0 * cm,
                2.5 * cm,
                2.7 * cm,
                2.5 * cm,
                2.3 * cm,
            ],
        )
        inst_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), primary_color),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
                    ("TOPPADDING", (0, 0), (-1, 0), 5),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_alt]),
                    ("GRID", (0, 0), (-1, -1), 0.5, border_color),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        story.append(inst_table)
        story.append(Spacer(1, 10))

        # 5. Contratos e Fornecedores
        story.append(Paragraph("Contratos & Fornecedores", section_style))
        contr_data = [
            [
                Paragraph("Fornecedor", cell_header_style),
                Paragraph("Valor Total", cell_header_style),
                Paragraph("Status do Contrato", cell_header_style),
            ]
        ]
        for contr in contracts:
            sup_name = contr.supplier.name if contr.supplier else "Fornecedor Direto"
            contr_data.append(
                [
                    Paragraph(sup_name, cell_style),
                    Paragraph(cls._format_currency(contr.total_amount), cell_style),
                    Paragraph(contr.get_status_display(), cell_style),
                ]
            )
        if len(contr_data) == 1:
            contr_data.append(
                [
                    Paragraph("Nenhum contrato cadastrado.", cell_style),
                    Paragraph("—", cell_style),
                    Paragraph("—", cell_style),
                ]
            )

        contr_table = Table(contr_data, colWidths=[8.0 * cm, 4.5 * cm, 4.5 * cm])
        contr_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), primary_color),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
                    ("TOPPADDING", (0, 0), (-1, 0), 5),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_alt]),
                    ("GRID", (0, 0), (-1, -1), 0.5, border_color),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        story.append(contr_table)
        story.append(Spacer(1, 10))

        # 6. Resumo do Checklist de Tarefas
        story.append(Paragraph("Checklist de Tarefas", section_style))
        tasks_completed = sum(1 for t in tasks if t.is_completed)
        tasks_total = len(tasks)
        tasks_pct = (
            round((tasks_completed / tasks_total) * 100) if tasks_total > 0 else 0
        )
        task_summary_text = (
            f"<b>Progresso:</b> {tasks_completed} de {tasks_total} "
            f"tarefas concluídas ({tasks_pct}%)"
        )
        story.append(Paragraph(task_summary_text, subtitle_style))

        # Renderização do documento
        doc.build(story, canvasmaker=NumberedCanvas)
        return buffer.getvalue()

    @classmethod
    def _build_excel_summary_sheet(
        cls,
        wb: Workbook,
        wedding: Any,
        overview: dict[str, Any],
        installments: list[Installment],
        tasks: list[Task],
    ) -> Worksheet:
        """Constrói e preenche a aba de Resumo Executivo da planilha."""
        ws_summary = cast(Worksheet, wb.active)
        ws_summary.title = "Resumo Executivo"

        title_font = Font(name="Calibri", size=14, bold=True, color="18181B")
        regular_font = Font(name="Calibri", size=11)

        now_label = datetime.now(UTC).strftime("%d/%m/%Y às %H:%M UTC")
        ws_summary.cell(
            row=1,
            column=1,
            value=f"Relatório: {wedding.groom_name} & {wedding.bride_name}",
        ).font = title_font
        ws_summary.cell(
            row=2,
            column=1,
            value=f"Emitido em: {now_label}",
        ).font = regular_font

        ws_summary.append([])
        ws_summary.append(["Informações Gerais", "Valor"])
        ws_summary.append(["Noivo", wedding.groom_name])
        ws_summary.append(["Noiva", wedding.bride_name])
        wedding_date_str = wedding.date.strftime("%d/%m/%Y") if wedding.date else "—"
        ws_summary.append(["Data do Casamento", wedding_date_str])
        ws_summary.append(["Local", wedding.location or "—"])
        ws_summary.append(["Convidados Estimados", wedding.expected_guests or "—"])
        ws_summary.append(["Status do Casamento", wedding.get_status_display()])

        ws_summary.append([])
        ws_summary.append(["Métrica Financeira", "Valor"])
        budget_obj = getattr(wedding, "budget", None)
        total_budget_val = float(budget_obj.total_estimated) if budget_obj else 0.0
        ws_summary.append(["Orçamento Total", total_budget_val])
        paid_sum = sum(
            (
                i.amount
                for i in installments
                if i.status == Installment.StatusChoices.PAID
            ),
            Decimal("0.00"),
        )
        ws_summary.append(["Total Pago", float(paid_sum)])

        pending_sum = sum(
            (
                i.amount
                for i in installments
                if i.status
                in (
                    Installment.StatusChoices.PENDING,
                    Installment.StatusChoices.OVERDUE,
                )
            ),
            Decimal("0.00"),
        )
        ws_summary.append(["Total Pendente / Atrasado", float(pending_sum)])

        budget_pct_used = overview.get("budget_percentage_used", 0)
        ws_summary.append(["Saúde Financeira Utilizada (%)", f"{budget_pct_used}%"])
        ws_summary.append(
            [
                "Tarefas Concluídas",
                f"{sum(1 for t in tasks if t.is_completed)} de {len(tasks)}",
            ]
        )
        return ws_summary

    @classmethod
    def _build_excel_categories_sheet(
        cls, wb: Workbook, categories: list[BudgetCategory]
    ) -> None:
        """Constrói a aba de Categorias Orçamentárias."""
        ws_cat = wb.create_sheet(title="Categorias Orçamentárias")
        ws_cat.append(
            [
                "Categoria",
                "Verba Alocada (R$)",
                "Total Gasto (R$)",
                "Saldo (R$)",
                "Uso (%)",
            ]
        )
        for cat in categories:
            spent = float(cat.total_spent)
            allocated = float(cat.allocated_budget)
            remaining = allocated - spent
            pct = round((spent / allocated) * 100, 1) if allocated > 0 else 0
            ws_cat.append([cat.name, allocated, spent, remaining, f"{pct}%"])

    @classmethod
    def _build_excel_installments_sheet(
        cls, wb: Workbook, installments: list[Installment]
    ) -> None:
        """Constrói a aba de Cronograma de Parcelas."""
        ws_inst = wb.create_sheet(title="Cronograma de Parcelas")
        ws_inst.append(
            [
                "Despesa / Item",
                "Parcela Nº",
                "Vencimento",
                "Valor (R$)",
                "Status",
                "Data de Pagamento",
            ]
        )
        for inst in installments:
            desc = inst.expense.description if inst.expense else "Parcela"
            due_str = inst.due_date.strftime("%d/%m/%Y")
            paid_str = inst.paid_date.strftime("%d/%m/%Y") if inst.paid_date else "—"
            ws_inst.append(
                [
                    desc,
                    inst.installment_number,
                    due_str,
                    float(inst.amount),
                    inst.get_status_display(),
                    paid_str,
                ]
            )

    @classmethod
    def _build_excel_contracts_sheet(
        cls, wb: Workbook, contracts: list[Contract]
    ) -> None:
        """Constrói a aba de Contratos & Fornecedores."""
        ws_contr = wb.create_sheet(title="Contratos & Fornecedores")
        ws_contr.append(
            [
                "Fornecedor",
                "Nome / Descrição",
                "Valor Total (R$)",
                "Status",
                "Data de Expiração",
            ]
        )
        for c in contracts:
            sup_name = c.supplier.name if c.supplier else "Fornecedor Direto"
            exp_str = (
                c.expiration_date.strftime("%d/%m/%Y") if c.expiration_date else "—"
            )
            ws_contr.append(
                [
                    sup_name,
                    c.name,
                    float(c.total_amount),
                    c.get_status_display(),
                    exp_str,
                ]
            )

    @classmethod
    def _build_excel_tasks_sheet(cls, wb: Workbook, tasks: list[Task]) -> None:
        """Constrói a aba de Checklist de Tarefas."""
        ws_tasks = wb.create_sheet(title="Checklist de Tarefas")
        ws_tasks.append(["Status", "Título da Tarefa", "Prazo", "Descrição"])
        for t in tasks:
            status_sym = "Concluída" if t.is_completed else "Pendente"
            due_str = t.due_date.strftime("%d/%m/%Y") if t.due_date else "—"
            ws_tasks.append([status_sym, t.title, due_str, t.description])

    @classmethod
    def _style_excel_workbook(cls, wb: Workbook, ws_summary: Worksheet) -> None:
        """Aplica bordas, cores de cabeçalho e largura automática nas colunas."""
        header_fill = PatternFill(
            start_color="0F766E", end_color="0F766E", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="E4E4E7"),
            right=Side(style="thin", color="E4E4E7"),
            top=Side(style="thin", color="E4E4E7"),
            bottom=Side(style="thin", color="E4E4E7"),
        )
        currency_format = '"R$" #,##0.00'
        monetary_keys = ["(r$)", "valor", "orçamento", "pago", "gasto", "saldo"]

        for sheet in wb.worksheets:
            for cell in sheet[1]:
                if sheet != ws_summary or cell.row in (4, 12):
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(cast(int, col[0].column))
                for cell in col:
                    cell.border = thin_border
                    col_idx = cast(int, cell.column)
                    header_val = str(sheet.cell(row=1, column=col_idx).value).lower()
                    if isinstance(cell.value, (int, float)) and any(
                        k in header_val for k in monetary_keys
                    ):
                        cell.number_format = currency_format
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    @classmethod
    def generate_wedding_excel(
        cls,
        company: Company,
        wedding_uuid: UUID | str,
    ) -> bytes:
        """
        Gera um relatório operacional completo do casamento em planilha Excel (.xlsx).

        Args:
            company: Empresa tenant autenticada.
            wedding_uuid: Identificador único do casamento.

        Returns:
            Bytes do arquivo Excel (.xlsx) gerado.
        """
        wedding = wedding_get_selector(
            company=company,
            uuid=UUID(str(wedding_uuid)),
        )
        overview = wedding_overview_selector(
            company=company,
            wedding_uuid=UUID(str(wedding_uuid)),
        )

        categories = list(
            BudgetCategory.objects.for_tenant(company)
            .filter(budget__wedding=wedding)
            .order_by("name")
        )
        installments = list(
            Installment.objects.for_tenant(company)
            .filter(wedding=wedding)
            .select_related("expense")
            .order_by("due_date")
        )
        contracts = list(
            Contract.objects.for_tenant(company)
            .filter(wedding=wedding)
            .select_related("supplier")
            .order_by("created_at")
        )
        tasks = list(
            Task.objects.for_tenant(company)
            .filter(wedding=wedding)
            .order_by("is_completed", "due_date")
        )

        wb = Workbook()
        ws_summary = cls._build_excel_summary_sheet(
            wb, wedding, overview, installments, tasks
        )
        cls._build_excel_categories_sheet(wb, categories)
        cls._build_excel_installments_sheet(wb, installments)
        cls._build_excel_contracts_sheet(wb, contracts)
        cls._build_excel_tasks_sheet(wb, tasks)
        cls._style_excel_workbook(wb, ws_summary)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @classmethod
    def generate_and_store_report(
        cls,
        company: Company,
        wedding_uuid: UUID | str,
        report_format: Literal["pdf", "excel"] = "pdf",
    ) -> tuple[str, str]:
        """
        Gera relatório, salva no Core Storage (R2/S3) e retorna URL pré-assinada.

        Args:
            company: Empresa tenant autenticada.
            wedding_uuid: Identificador único do casamento.
            report_format: Formato desejado ('pdf' ou 'excel').

        Returns:
            Tupla (object_key, presigned_download_url).
        """
        wedding = wedding_get_selector(
            company=company,
            uuid=UUID(str(wedding_uuid)),
        )

        now_ts = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
        if report_format == "excel":
            file_bytes = cls.generate_wedding_excel(company, wedding.uuid)
            content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            extension = "xlsx"
        else:
            file_bytes = cls.generate_wedding_pdf(company, wedding.uuid)
            content_type = "application/pdf"
            extension = "pdf"

        bucket = getattr(settings, "AWS_STORAGE_BUCKET_NAME", "wedding-reports")
        company_uuid = company.uuid
        wedding_uuid_val = wedding.uuid
        object_key = (
            f"reports/{company_uuid}/{wedding_uuid_val}/relatorio_{now_ts}.{extension}"
        )

        storage = cls._get_storage_service()
        saved_key = storage.upload_bytes(
            bucket=bucket,
            object_key=object_key,
            data=file_bytes,
            content_type=content_type,
        )
        download_url = storage.generate_presigned_get_url(
            bucket=bucket,
            object_key=saved_key,
            expires_in=3600,  # 1 hora de validade
        )
        return saved_key, download_url
